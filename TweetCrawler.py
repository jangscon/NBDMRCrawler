from NBDMRCrawler import NBDMRCrawler
import pkg_resources

class TweetCrawler(NBDMRCrawler):
    '''
    tweet crawler class
    super : class NBDMRCrawler
    '''

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        # TODO put brands in self.brand_names
        self.brand_names = {}

        # attributes for twitter crawling
        self.search_url = "https://api.twitter.com/2/tweets/search/all"
        self.query_params = {'query': '', 'tweet.fields': 'created_at,lang,author_id',
                             'start_time': '2016-01-01T00:00:00Z',
                             'end_time': '2021-01-01T00:00:00Z', 'max_results': '500' }
        self.doccano_threshold = float(self.doccano_threshold)
        self.mongodb_access = self.get_mongoDB_access(host=self.mongodb_host, db_name=self.mongodb_db , collection_name=self.mongodb_col)

        self.doccano_access = self.get_doccano_access( baseurl=self.doccano_url , username=self.doccano_username , password=self.doccano_password)


    '''
    setting brands attribute
    '''

    def get_brands_find_args_with_fisrtchar(self , char_list):
        regex_str = f"({'|'.join([f'^{i}'  for i in char_list])})"
        return {"drug": {"$regex": regex_str, "$options": 'i'}}

    def set_brands_RESTAPI(self, islist=False):
        '''
        get_product_name method의 return 값을 brand_names attribute에 update하는 method

        :param islist: type( Boolean ) - list or dict
        :return: None - setting brands_names attribute
        '''

        brands = self.get_product_name(druglist=self.drugs, type_var=islist)
        self.brand_names = brands

    def set_drugs_brands_from_mongoDB(self ,mongoDB_col ,
                                 find_args={},
                                 find_kwargs={"_id":0 , "update_time":0},
                                 ):
        import pymongo

        drugs_brands = mongoDB_col.find(find_args,find_kwargs)
        drugs = []
        brands = {}
        for i in drugs_brands :
            drugs.append(i['drug'])
            brands.update({i['drug'] : i['brands']})
        d_len = len(drugs)
        b_len = sum([len(brands[x]) for x in brands if isinstance(brands[x], list)])
        print(f"drugs count : {d_len}")
        print(f"brands count : {b_len}")
        print(f"totla count : {d_len + b_len}")
        self.drugs = drugs
        self.brand_names = brands
        self.total_count = d_len + b_len

    '''

    crawling data processing

    '''

    def get_patterns_to_preprocessing(self , pattern_keywords : list):
        import re

        pattern_rt = re.compile(r'@RT[\S\s]+')
        pattern_special = re.compile(r'[=+#:^$*\"※~&%ㆍ』\\‘|\[\]`…》]')
        pattern_email = re.compile(r"\S*@\S*\s?")
        pattern_link = re.compile(r'http\S+')
        pattern_consecutive = re.compile(r"([a-zA-Z])\1{2,}")
        repl_consecutive = r"\1\1"
        pattern_colon = re.compile(r'^:')
        pattern_spclcodes = re.compile(r'(&nbsp;){,}(&lt;){,}(&gt;){,}(&amp;){,}(&quot;){,}(&#035;){,}(&#039;){,}')
        pattern_brasket = re.compile(r'(\[\]){,}')

        patterns = {
            "RT": pattern_rt,
            "SPCL": pattern_special,
            "EMAIL": pattern_email,
            "LINK": (pattern_link, "__url__"),
            "CONSCTV": (pattern_consecutive, repl_consecutive),
            "COLON": pattern_colon,
            "HTMLSPCL": pattern_spclcodes,
            "BRASKET": pattern_brasket,
        }

        result = []

        for keyword in pattern_keywords:
            if keyword in patterns:
                result.append(patterns[keyword])
            else:
                pass

        return result

    def tweet_preprocessing(self, text, patterns=[]):

        import re
        import demoji
        import preprocessor as p

        # using preprocessor module
        text = str(text)

        # using regex
        if patterns:
            for pattern in patterns:
                if isinstance(pattern, tuple):
                    text = re.sub(pattern=pattern[0], repl=pattern[1], string=text)
                else:
                    text = re.sub(pattern=pattern, repl='', string=text)
        else:
            pass

        p.set_options(p.OPT.EMOJI, p.OPT.SMILEY, p.OPT.MENTION, p.OPT.RESERVED, p.OPT.MENTION)
        text = p.clean(text)

        return text

    def post_processing(self , tweet , sui_classifier , regex_patterns):
        import preprocessor as p

        tw_obj = tweet.get("tw_obj")

        sui_keywords = self.s2s_sui_keywords
        warnings = self.trigger_warnings

        if tw_obj is not None and tw_obj["text"]:

            data = {
                "tweet_id": int(tw_obj["id"]) if tw_obj["id"].isdigit() else tw_obj["id"],
                "lang": tw_obj["lang"],
                "raw_text" : tw_obj["text"] ,
                "text": tw_obj["text"],
                "is_retweet" : True if tw_obj["text"][:3] == "RT " else False,
                "warnings" : [],
                "sui_keywords" : [],
                "author_id": tw_obj["author_id"],
                "drugs": [tweet.get("drug")],
                "created_at": tw_obj["created_at"]
            }

            check_hashtag = str(data["text"])
            p.set_options(p.OPT.URL , p.OPT.HASHTAG)
            check_hashtag = p.clean(check_hashtag)
            check_hashtag = check_hashtag.strip()

            if len(check_hashtag) < 20 :
                return None 
 
            data["text"] = self.tweet_preprocessing(text=data["text"], patterns=regex_patterns) 
            # sui_classifier
            if data["text"] is None or data["text"] == "" : 
                return None
            else :
                k = sui_classifier(data["text"], self.s2s_candidate_labels)
                data.update({"sui_classifier": k.get("scores")[k.get("labels").index("suicide")]})

            # trigger warnings
            wrn_minidx = 999
            tw = []

            for wrn in warnings :
                wrn_idx = data["text"].find(wrn)
                if wrn in data["text"] and wrn_idx < wrn_minidx :
                    wrn_minidx = wrn_idx
                    wrn_txt = data["text"][ wrn_minidx + len(wrn) : ]
                    wrn_txt = wrn_txt.split('\n')[0].strip()
                    tw = list(map(lambda s : s.strip() , wrn_txt.split(',')))
                else : pass
            data["warnings"] = tw

            # sui_keywords
            if sui_keywords:
                suiskw = [kw for kw in sui_keywords if str(data["text"].lower()).find(kw.lower()) != -1]
                data.update({"sui_keywords": suiskw})
            else:
                pass


            no_drugs_in_text = True
            for a in data["drugs"]:
                if a.lower() in data["text"].lower():
                    no_drugs_in_text = False
                else : pass

            if no_drugs_in_text:
                return None
            else : pass


            if data is None:
                return None
            else : pass
           
            return data

    def distance_btw_two_str(self, a, b):
        from difflib import SequenceMatcher
        return SequenceMatcher(None, a.lower(), b.lower()).ratio()

    def get_tokenizer(self):
        import nltk.data
        import warnings

        warnings.filterwarnings('ignore')
        tokenizer = nltk.data.load('tokenizers/punkt/english.pickle')

        return tokenizer

    def get_sui_clasifier(self):
        from transformers import pipeline
        import warnings

        warnings.filterwarnings('ignore')
        sui_classifier = pipeline("zero-shot-classification", model="facebook/bart-large-mnli")


        return sui_classifier

    def get_med7_model(self):
        import spacy
        import warnings

        warnings.filterwarnings('ignore')
        med7_model = spacy.load("en_core_med7_trf")

        return med7_model

    def split_to_sents(self, tweet,tokenizer , sui_classifier , med7_model ):


        if tweet is None :
            return None
        else :
            pass

        sui_keywords = self.s2s_sui_keywords
        med7_distance = self.s2s_med7_distance
        med7_looptime = self.s2s_med7_looptime
        exc_sents = self.s2s_exc_sents
        candidate_labels = self.s2s_candidate_labels



        ss = tokenizer.tokenize(tweet['text'])
        tlist = []

        # tweet

        # sui clasifier update
        k = sui_classifier(tweet["text"], candidate_labels)
        tweet.update({"sui_clasifier": k.get("scores")[k.get("labels").index("suicide")]})

        # sui keywords update
        if sui_keywords:
            suiskw = [kw for kw in sui_keywords if kw.lower() in tweet["text"].lower()]
            tweet.update({"sui_keywords": suiskw})
        else:
            tweet.update({"sui_keywords": []})

        # sentences
        for s in ss:
            if s not in exc_sents:
                sdict = {
                    "sent": s,
                    "drugs": [drug for drug in tweet['drugs'] if drug.lower() in s.lower()]
                    # if isinstance(tweet['drugs'], list) else [tweet['drugs']] if tweet['drugs'].lower() in s.lower() else [],

                }
                # if sdict["drugs"] == [] :
                #     print(f"tweet drugs : {tweet['drugs']}" )
                #     print(f"sents : {sdict['sent']}")

                # sdict["brands"] = [tweet['brand'][i] for i, v in enumerate(tweet["drugs"]) if
                #                    v in sdict["drugs"]] if sdict["drugs"] else []

                if sui_keywords:
                    suiskw = [kw for kw in sui_keywords if kw.lower() in s.lower()]
                    sdict.update({"sui_keywords": suiskw})
                else :
                    sdict.update({"sui_keywords": []})

                '''
                classifier
                '''

                k = sui_classifier(s, candidate_labels)
                sdict.update(
                    {"sui_clasifier": k.get("scores")[k.get("labels").index("suicide")]})  # score for suicide
                # sdict.update({"sui_classifier": k.get("labels")[k.get("scores").index(max(k.get("scores")))]}) #a label

                '''
                med_7
                '''

                med7_loop = med7_looptime
                sdict.update({"med7": []})
                # med7 is not working ex:api problem - try few times
                for i in range(med7_loop) :
                    med7drugs = [{"text": ent.text, "start_char": ent.start_char, "end_char": ent.end_char , "distance":[]} for ent
                                in med7_model(s).ents if ent.label_ == "DRUG"]
                    if med7drugs:
                        sdict.update({"med7": med7drugs})
                        break
                    else : pass


                # sdict dict has med7 key -> drugs key's values and med7.text key's values are compared through distance,
                # and distance is greater than the limit and drugs key's value, which is the largest, is determined again as drups.
                # TODO drugs가 빈 칸으로 나오는 경우
                if "med7" in sdict :
                    new_drugs = []
                    for d in sdict["drugs"] :
                        distance = 0
                        for med7_d in sdict["med7"] :
                            result_distance = self.distance_btw_two_str(med7_d["text"] , d)
                            med7_d["distance"].append({d : result_distance})
                            distance = result_distance if distance < result_distance else distance
                        if distance >= float(med7_distance) :
                            new_drugs.append(d)
                        else : pass
                    sdict["drugs"] = new_drugs
                else : pass

                # if sdict["drugs"] and "med7" in sdict :
                #     tlist.append(sdict)
                tlist.append(sdict)

            if tlist :
                tweet.update({"sentences": tlist})

        return tweet

    '''

    crawling part 

    '''

    def create_headers(self):
        '''
        인증을 위한 http header 정보를 가지는 dict를 반환하는 함수

        :return: type(dict) - http header
        '''
        headers = {"Authorization": "Bearer {}".format(self.bearer_token)}
        return headers

    def crawling_part(self, drug_name):
        '''
        connect_to_endpoint를 통해 얻은 json_response를 results attribute에 저장하는 method

        재귀 호출 사용

        :param brand_list: type( list ) - brand name list to drug_name
        :param drug_name: type( str ) - drug name string
        :return: None - self.crawling_part(drug_name , brand_name)
        '''

        import time
        import json

        headers = self.create_headers()
        # json_response = self.connect_to_endpoint(headers)

        #f = open("rqst.txt", 'a')
        #now = time.time()

        json_response = self.connect_to_endpoint_with_retry(headers=headers)

        #stop = str(time.time() - now)
        #f.writelines(stop + "\n")
        #f.close()

        # data = json.dumps(json_response, indent=4, sort_keys=True, ensure_ascii=False).encode('utf8')
        # results update

        if json_response:  # json_response를 정상적으로 가져오고 - HTTP GET 200
            if "data" in json_response:  # data라는 key를 가지면
                for i in json_response["data"]:
                        yield {"drug": drug_name, "tw_obj": i}
            else : pass
        else : pass
        
        try:
            self.query_params['next_token'] = json_response['meta']['next_token']
        except Exception as e:
            print("\t\t- " + self.query_params['query'] + " done")
            self.query_params['next_token'] = "DUMMY"
            del self.query_params['next_token']
            return None

        self.crawling_part(drug_name )

    def crawling_part_id(self):
        import json
        import time
        print(self.query_params_id)
        headers = self.create_headers()
        json_response = self.connect_to_endpoint_id(headers)
        data = json.dumps(json_response, indent=4, sort_keys=True)
        try:
            self.query_params_id['pagination_token'] = json_response['meta']['next_token']
        except Exception as e:
            print("done")
            self.query_params_id['pagination_token'] = "DUMMY"
            del self.query_params_id['pagination_token']
            # print(e)
            time.sleep(2)
            return
        time.sleep(2)
        self.crawling_part_id()

    def main_act_crawling(self,drug_name ,  from_date, to_date , sui_classifier, pp_regex_patterns, tokenizer=None, med7_model=None ):
        from tqdm import tqdm
        import time
        from MongoDB_duplicate_handling import delete_duplicate_data_in_mongoDB

        self.query_params['start_time'] = str(from_date) + "T00:00:01Z"
        self.query_params['end_time'] = str(to_date) + "T00:00:00Z"
        self.query_params['query'] = f"\"{drug_name}\" {self.tweet_filter_query}"


        for result in self.crawling_part(drug_name) :
            if result :
                tweet = self.post_processing(result ,
                                             sui_classifier=sui_classifier ,
                                             regex_patterns=pp_regex_patterns)
                if tweet is None :
                    continue

                # print(f"\t\tsplit to sents - {drug_name}...   ", end='\r')

                # tweet = self.split_to_sents(tweet=tweet ,
                #                             tokenizer=tokenizer,
                #                             sui_classifier=sui_classifier,
                #                             med7_model=med7_model
                #                             )

                # print(f"\t\tsplit to sents - {drug_name} done", end='\r')
                # print(f"\t\tupdate mongodb - {drug_name}...   ", end='\r')
                try :
                    self.get_and_update_tweet_mongoDB(mongDB_collection=self.mongodb_access,
                                                  update_tweet=tweet)
                    delete_duplicate_data_in_mongoDB(self.mongodb_access , "text")
                except Exception as e :
                    print("mongo exception")
                    self.mongodb_access = self.get_mongoDB_access(host=self.mongodb_host, db_name=self.mongodb_db,
                                                                  collection_name=self.mongodb_col)
                    pass
                

                try :
                    self.import_dict_to_doccano(doccano_client=self.doccano_access,
                                            project_id=self.doccano_pid,
                                            import_dict=self.get_dict_to_import_doccano(tweet),
                                            )
                except Exception as e :
                    print("doccano exception")
                    self.doccano_access = self.get_doccano_access(baseurl=self.doccano_url,
                                                                  username=self.doccano_username,
                                                                  password=self.doccano_password)
                    pass


                
                # print(f"\t\tupdate mongodb - {drug_name} done " ,end='\r')
            else : pass
    def get_donedrugs_list(self , config_path):
        import pickle
        
        with open(config_path, 'rb') as lf:
            readList = pickle.load(lf)
        return readList

    def update_donedrugs(self ,drug_list, drug_name , config_path):
        import pickle

        with open(config_path, 'wb') as lf:
            pickle.dump(drug_list+[drug_name], lf)

        return self.get_donedrugs_list(config_path)


    def crawling_by_brand_and_drug(self):
        from_date, to_date = self.make_datetime()
        done_drugs = self.get_donedrugs_list(config_path="config/donedrugs.txt")
        total_drugs_brands =  list(set(sum(list(self.brand_names.values()), []) + self.drugs) - set(done_drugs))
        print(f"total drugs & brands length : {len(total_drugs_brands)}")
        print(f"done_drugs length : {len(list(set(done_drugs)))}")
        print("load split to sents paramters...",end="\r")
        # tokenizer = self.get_tokenizer()
        sui_clsfr = self.get_sui_clasifier()
        # med7_model = self.get_med7_model()
        pp_regex_patterns = self.get_patterns_to_preprocessing(self.regex_pattern_keywords)

        print("crawling total drugs & brands")
        for i,drug in enumerate(total_drugs_brands):
            print(f"NO.{i+1} ({((i+1)/self.total_count):.4f}%) - ",end='')
            self.main_act_crawling(drug_name=drug, from_date=from_date, to_date=to_date ,
                                   sui_classifier=sui_clsfr ,
                                   pp_regex_patterns=pp_regex_patterns)
            done_drugs = self.update_donedrugs(drug_list=done_drugs , drug_name=drug ,config_path="config/donedrugs.txt")
            


    '''

    mongoDB

    '''

    def export_txt_from_url(self, host, collection_name, documents_name, source="tweet", search_q=None, file_name=None,
                            text_width=50):
        '''
        몽고DB의 데이터를 텍스트 파일로 저장하는 method입니다.

        @param host: type(string) - 주소와 비밀번호를 포함한 문자열입니다.
        @param collection_name: type(string) - 컬렉션의 이름 문자열입니다.
        @param documents_name: type(string) - 컬렉션 내부의 특정 도큐먼트 이름 문자열입니다.
        @param source: type(str) - 몽고디비 데이터의 출처입니다.
        @param search_q: type(dict) - 몽고디비의 특정 데이터를 찾는 쿼리문입니다.
        @param file_name: type(str) - 저장할 텍스트 파일의 이름입니다.
        @param text_width: type(int) - 텍스트를 write 시에 간격입니다. 초기값 50이면 50자를 쓰고 줄바꿈합니다.
        @return: None = write & save txt file
        '''
        import pymongo
        from datetime import datetime
        import pkg_resources

        required = {'pyfiglet'}
        installed = {pkg.key for pkg in pkg_resources.working_set}
        missing = required - installed

        if not file_name:
            now = datetime.now()
            nowdate = now.strftime("%Y%m%d")
            file_name = "{}{}_urls.txt".format(source, nowdate)
        else : pass
        
        # 아스키아트 출력 - pyfiglet module이 없으면 실행 안함
        if not missing:
            # 아스키아트 부분 - 필요없으니 삭제해도 상관 없습니다!
            from pyfiglet import Figlet
            f = Figlet(font='slant')
            print(f.renderText('-----------'))
            print(f.renderText('Mongo to Txt'))
            print(f.renderText('-----------'))
            print("KNU DEAL Lab - {}\n".format(now.strftime("%Y/%m/%d")))
        else : pass
        
        try:
            client = pymongo.MongoClient(host)
        except Exception as e:
            print(e)
            print('\033[31m' + "Invalid MongoDB Client Connection string" + '\033[0m')
            return
            
        try:
            db = client[collection_name]
        except Exception as e:
            print(e)
            print('\033[31m' + "Invalid MongoDB collection name" + '\033[0m')
            return
            
        try:
            col = db[documents_name]
        except Exception as e:
            print(e)
            print('\033[31m' + "Invalid MongoDB document name" + '\033[0m')
            return

        cursor = col.find() if search_q == None else col.find(search_q)
        db_collections = list(cursor)

        print(f"\nwrite {file_name}")
        with open(file_name, "w") as f:
            for i, v in enumerate(db_collections):
                text = []
                for j, k in enumerate(v['text']):
                    text.append(k)
                    if (j + 1) % text_width == 0:
                        text.append("\n  ")
                    else : pass
                    
                text = "".join(text)
                f.write(f"- {text}\n\n")
                f.write(f"\turl : {'https://twitter.com/anyuser/status/' + v['_id']}\n\n\n")
        print(f"save {file_name}")

    def advance_cell_size(self, file_name):
        '''
        엑셀파일을 load하여 셀의 column과 row를 알맞게 조절한 후 수정된 엑셀파일을 저장하는 메서드

        :param file_name:  type(string) - 불러올 엑셀파일 이름입니다.
        :return: none - 엑셀파일을 수정 후 저장합니다.
        '''
        import openpyxl
        print("\nfix cells Width & Height")

        wb = openpyxl.load_workbook(filename=file_name, keep_vba=False)
        sheet = wb['label']
        # len(sentences_df) + 2
        for i in range(2, sheet.max_row + 1):
            sheet.row_dimensions[i].height = 20
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 20

        sheet = wb['result']
        for i in range(1, sheet.max_row + 1):
            sheet.row_dimensions[i].height = 20
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 45
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 30

        wb.save(file_name)
        print("fix cells Width & Height done")

    def export_excelfile(self, host, collection_name, documents_name, limit=None, search_q=None, file_name=None,
                         sentences_keys=None):
        '''
        몽고DB에서 가져온 데이터를 엑셀파일로 만드는 메서드 입니다.

        :param host: type(string) - 주소와 비밀번호를 포함한 문자열입니다.
        :param collection_name: type(string) - 컬렉션의 이름 문자열입니다.
        :param documents_name: type(string) - 컬렉션 내부의 특정 도큐먼트 이름 문자열입니다.
        :param limit: type(int) - 엑셀파일에 담을 데이터의 수를 제한하는 정수입니다. 초기값 None은 제한이 없이 모든 데이터를 export합니다.
        :param search_Q: type(dict) - 딕셔너리 형태의 쿼리문 입니다. 초기값 None은 모든 데이터를 MongoDB에서 가져옵니다.
        :param file_name: type(string) - 저장할 엑셀파일 이름입니다. 초기값 None을 사용시 Tweet20220209_to_label.xlsx 과 같이 그날의 날짜에 맞춰서 파일이 생성됩니다.
        :param sentences_keys: type(list) - sentence sheet에 추가할 column들의 list 입니다. 사용자 입력은 받지 않습니다.
        :return: None - 엑셀 파일이 저장됩니다.
        '''
        import pymongo
        from styleframe import StyleFrame
        from styleframe import Styler
        import pandas as pd
        from tqdm import tqdm
        import pkg_resources
        import os.path
        from datetime import datetime

        now = datetime.now()
        r_col = ['Mongo_id', 'Tweet_id', 'lang', 'text', 'author_id', 'drug', 'brand', 'created_at']
        s_col = ['sentence', 'Drug_label', 'Brand_label']

        if sentences_keys:
            s_col.extend(sentences_keys)
        else : pass
        
        required = {'pyfiglet'}
        installed = {pkg.key for pkg in pkg_resources.working_set}
        missing = required - installed

        if not file_name:
            nowdate = now.strftime("%Y%m%d")
            file_name = "labels/{}{}_to_label_1.xlsx".format(documents_name, nowdate)
            file_num = 1
            while True:
                if os.path.isfile(file_name):
                    file_num += 1
                    file_name = "labels/{}{}_to_label_{}.xlsx".format(documents_name, nowdate, file_num)
                    continue
                break
        else:
            file_name = "labels/" + file_name

        # 아스키아트 출력 - pyfiglet module이 없으면 실행 안함
        if not missing:
            # 아스키아트 부분 - 필요없으니 삭제해도 상관 없습니다!
            from pyfiglet import Figlet
            f = Figlet(font='slant')
            print(f.renderText('\n\n-------------'))
            print(f.renderText('Mongo to Excel'))
            print(f.renderText('-------------'))
            print("KNU DEAL Lab - {}\n".format(now.strftime("%Y/%m/%d")))

        results_df = pd.DataFrame(columns=r_col)
        sentences_df = pd.DataFrame(columns=s_col)
        sentences = []

        try:
            client = pymongo.MongoClient(host)
        except Exception as e:
            print(e)
            print('\033[31m' + "Invalid MongoDB Client Connection string" + '\033[0m')
            return
        try:
            db = client[collection_name]
        except Exception as e:
            print(e)
            print('\033[31m' + "Invalid MongoDB collection name" + '\033[0m')
            return
        try:
            col = db[documents_name]
        except Exception as e:
            print(e)
            print('\033[31m' + "Invalid MongoDB document name" + '\033[0m')
            return

        cursor = col.find() if search_q == None else col.find(search_q)
        db_collections = list(cursor)
        print("MongoDB : {}{}/{}".format(host, collection_name, documents_name))
        print("total docs in collection: {}".format(col.count_documents({})))
        print("total docs returned by find() with Query: {}\n".format(len(db_collections)))

        # db 저장된 tweet data를 dataframe으로 변환
        for item in tqdm(db_collections[:limit], desc="db_collections to Dataframe..."):
            temp_append = []
            for key, value in item.items():
                if key == "sentences":
                    for i in value:
                        if i['brands']:
                            temp_dict = {"id": item["tweet_id"]}
                            temp_dict.update(i)
                            sentences.append(list(temp_dict.values()))
                        else : pass
                else:
                    temp_append.append(value)

            # 5 : drugs
            # 6 : brand

            for idx, v in enumerate(temp_append[5]):
                ra = temp_append[:]
                ra[5] = v
                ra[6] = ra[6][idx]
                results_df = results_df.append(pd.Series(ra, index=results_df.columns), ignore_index=True)

        from operator import itemgetter
        sorted_sentences = sorted(sentences, key=itemgetter(0))

        # tweet data에서 sentence 부분을 dataframe으로 변환
        for sentence in tqdm(sorted_sentences, desc="sentences data to Dataframe..."):
            for idx, v in enumerate(sentence[4]):
                s = sentence[:]
                try:
                    s[3] = s[3][idx] if s[3] else ""
                except Exception as e:
                    pass

                s[4] = v
                s = s[2:]
                sentences_df = sentences_df.append(pd.Series(s, index=sentences_df.columns), ignore_index=True)

        # header 와 column의 styler
        header_styler = Styler(bold=True, horizontal_alignment="center")
        column_styler = Styler(horizontal_alignment="left")

        # dataframe들을 styleframe으로 변환 후 style을 입혀서 엑셀 파일로 저장
        print("\nDataframe to Excel file...")
        writer = StyleFrame.ExcelWriter(file_name)
        sf = StyleFrame(results_df)
        sf.apply_headers_style(styler_obj=Styler.combine(header_styler, Styler(bg_color="yellow")))
        sf.apply_column_style(cols_to_style=r_col, styler_obj=column_styler)
        sf.to_excel(writer, index=False, best_fit=r_col, sheet_name="result", columns_and_rows_to_freeze='B2',
                    row_to_add_filters=0)
        sf = StyleFrame(sentences_df)
        sf.apply_headers_style(styler_obj=Styler.combine(header_styler, Styler(bg_color="green")))
        sf.apply_column_style(cols_to_style=s_col, styler_obj=column_styler)
        sf.to_excel(writer, index=False, best_fit=s_col, sheet_name="label", columns_and_rows_to_freeze='B2',
                    row_to_add_filters=0)
        writer.save()
        print("{} saved".format(file_name))

        # column 간 간격을 조정하는 method
        self.advance_cell_size(file_name)

    def get_and_update_tweet_mongoDB(self , mongDB_collection, update_tweet, no_s2s=True):
        '''

        tweet (dict) -
            drugs (list) - update
            sentences (list) -
                            drugs (list) - update
                            distance (list) -
                                            distance(list) - update

        '''
        from operator import itemgetter

        if update_tweet is not None:
            mongodata = mongDB_collection.find_one({"tweet_id": update_tweet['tweet_id']})

            if mongodata is None:
                mongDB_collection.insert_one(update_tweet)
            else:
                # tweet.drugs update
                update_tweet["drugs"] = list(set(update_tweet["drugs"] + mongodata["drugs"]))

                if no_s2s is False:
                    # tweet.sentences.drugs update
                    for sent, update_sent in zip(mongodata["sentences"], update_tweet["sentences"]):
                        update_sent["drugs"] = list(set(sent["drugs"] + update_sent["drugs"]))

                        # tweet.sentences.med7 update
                        if sent["med7"]:
                            sent["med7"] = sorted(sent["med7"], key=itemgetter("text"))
                            update_sent["med7"] = sorted(update_sent["med7"], key=itemgetter("text"))
                            for med7text, update_med7text in zip(sent["med7"], update_sent["med7"]):
                                ismed7 = [k for dist in update_med7text["distance"] for k, i in dist.items()]
                                temp_dist = []
                                for i in med7text["distance"]:
                                    if list(i.keys())[0] in ismed7:
                                        temp_dist.append(i)
                                    else:
                                        pass
                                update_med7text["distance"].extend(temp_dist)
                        else:
                            pass
                else:
                    pass
                    # update tweet
                mongDB_collection.update_one({"tweet_id": update_tweet['tweet_id']}, {"$set": update_tweet},
                                             upsert=True)
        else:
            pass


    '''
    
    doccano
    
    '''

    def get_doccano_access(self , baseurl: str, username: str, password: str):
        from doccano_api_client import DoccanoClient

        doccano_client = DoccanoClient(
            baseurl,
            username,
            password
        )
        return doccano_client

    def get_dict_to_import_doccano(self, tweet : dict):
        tweet_keys = ["text","tweet_id","author_id", "sui_classifier" , "drugs", "sui_keywords" , "warnings"]
        result = { twk : tweet[twk] for twk in tweet_keys if tweet[twk] }
        result["tweet_id"] = str(result["tweet_id"])
        return result

    def import_dict_to_doccano(self , doccano_client, project_id, import_dict,
                               threshold_field="sui_classifier", column_data="text", column_label="label"):
        from doccano_api_client import DoccanoClient
        from io import BytesIO, TextIOWrapper

        if "sui_keywords" in import_dict or "warnings" in import_dict or import_dict[threshold_field] >= self.doccano_threshold : 
            bio = BytesIO()
            bio.write(json.dumps(import_dict).encode())
            bio.seek(0)
            with TextIOWrapper(bio, encoding="utf-8") as f:
                doccano_client.post_doc_upload_binary(
                    project_id=project_id,
                    files=[f],
                    column_data=column_data,
                    column_label=column_label,
                    delimiter="",
                    encoding="utf_8",
                    format="JSONL",
                )
        else :
            pass

    '''

    make histogram

    '''

    def make_histogram_data_and_StatisticsDict(self , biger_suiclsf=None):
        '''
        tweets.sentences의 sui_classifier value들로 만든 sorted list와 통계적인 값들로 만든 dictionary로 된 tuple을 반환하는 method

        @biger_suiclsf : type(float) - append greater numbers than biger_suiclsf into the sui_clasifier_list.
        @return: type(typle(list , dict)) - return tuple composed of list of sui_clasifier values and statistics data dictionary.

        '''
        import numpy as np

        sui_clasifier_list = []
        statistics_dict = {}

        for i in self.tweets:
            if "sentences" in i:
                for j in i["sentences"]:
                    if biger_suiclsf == None  :
                        sui_clasifier_list.append(j["sui_clasifier"])
                    else :
                        if j["sui_clasifier"] >= biger_suiclsf :
                            sui_clasifier_list.append(j["sui_clasifier"])
                        else :
                            pass
            else :
                pass

        # sui_classifier list sorting
        sui_clasifier_list = sorted(sui_clasifier_list)

        # make statisitics data
        statistics_dict["total_num"] = len(sui_clasifier_list)
        statistics_dict["min_num"] = sui_clasifier_list[0]
        statistics_dict["max_num"] = sui_clasifier_list[-1]
        statistics_dict["mean"] = np.mean(sui_clasifier_list)
        statistics_dict["variance"] = np.var(sui_clasifier_list)
        statistics_dict["std"] = np.std(sui_clasifier_list)

        return (sui_clasifier_list, statistics_dict)

    def make_crawling_inform(self):
        '''
        크롤링관련 저장할 데이터를 문자열로 반환하는 method

        @return: type(str) - related crawler data string
        '''
        result = {}
        result["source"] = f"{self.source}"
        result["date"] = f"{self.from_date} ~ {self.to_date}"

        return result

    def make_histogram(self, datalist, crawling_inform, statistics_data, file_name):
        '''
        히스트그램을 만드는 method

        @param datalist: type(list) - data to make a histogram
        @param crawling_inform: type(str) - string with crawler infromation
        @param statistics_data: type(dict) - dictionary with statistics data
        @param file_name: type(str) - file name to save histogram file
        @return: None - make & save histogram file ( .png )
        '''
        import matplotlib.pyplot as plt
        import matplotlib.offsetbox as offsetbox
        plt.figure(figsize=(20, 10))
        txt = " "
        for k, v in crawling_inform.items():
            if txt[-1] == ",":
                txt += f" {k} : {v}\n"
            else:
                txt += f"{k} : {v} ,"
        if txt[-1] == ",":
            txt = txt[:-1]
            txt += "\n"
        for k, v in statistics_data.items():
            txt += f"{k} : {v}\n"
        plt.hist(datalist, edgecolor='black', linewidth=1.2, bins=100)
        textbox = offsetbox.AnchoredText(txt, loc=1)
        plt.gca().add_artist(textbox)
        plt.xlabel('num', labelpad=15)
        plt.ylabel('count', labelpad=10)
        plt.xlim([0, 1])
        plt.xticks([0.05 * x for x in range(0, 20)])
        plt.savefig(file_name)

# user's prompt print with color
def print_with_color( text, tfs_color):
    print(f'\033[38;5;{tfs_color}m' + text + '\033[0m')


def main(**kwargs):

    # crawlerObj에서 필요 없는건 pop 후 저장 - histogram , excel export
    histogram_name = kwargs.pop("histogram_name")
    is_excel_save = kwargs.pop("is_excel_save")
    excelDB_limit = kwargs.pop("excel_limit")
    excelDB_query = kwargs.pop("excel_query")
    excelDB_filename = kwargs.pop("excel_name")
    excelDB_sentences_keys = kwargs.pop("sentences_keys")
    histogram_sui_clasifier_limit = kwargs.pop("histogram_sui_clasifier_limit")

    # crawler version
    crawler_version = kwargs.pop("version")

    # ascii art
    required = {'pyfiglet'}
    installed = {pkg.key for pkg in pkg_resources.working_set}
    missing = required - installed

    # pyfiglet modlue이 있으면 실행되는 코드
    if not missing:
        # 아스키아트 부분 - 필요없으니 삭제해도 상관 없습니다!
        from pyfiglet import Figlet
        f = Figlet(font='slant')
        print(f.renderText('-------------'))
        print(f.renderText(' Tweet Crawler'))
        print(f.renderText('-------------'))
        print(f"NBDMRcrawler_{crawler_version} - Tweet Crawler ")
    else : pass

    # crawler obj 생성
    crawlerObj = TweetCrawler(**kwargs)

    '''

        parsing argument Exception check 


    '''

    # argparser로 from date를 받지 못했을 경우
    from_date = kwargs["from_date"]
    if not crawlerObj.check_from_date(from_date, first_check=True):

        # from_date 사용자 입력
        print_with_color("\n1-1. Enter the start date of the data to crawling.", 33)
        print_with_color("     input form - (ex : 2022/03/01)", 33)
        while True:
            # 사용자 입력
            from_date = input(":")
            if not crawlerObj.check_from_date(from_date):
                continue

            # crawling obj의 from_date attribute로 update
            print_with_color(f"\n\tfrom_date : {from_date}\n", 28)
            crawlerObj.from_date = from_date
            break
    else : pass

    # argparser로 to date를 받지 못했을 경우
    to_date = kwargs["to_date"]
    if not crawlerObj.check_to_date(to_date, from_date, first_check=True):
        # to_date 사용자 입력
        print_with_color("\n1-2. Enter the end date of the data to crawling.", 33)
        print_with_color("     input form - (ex : 2022/03/02)", 33)
        while True:
            # 사용자 입력
            to_date = input(":")
            if not crawlerObj.check_to_date(to_date, from_date):
                continue
            else : pass

            # crawling obj의 to_date attribute로 update
            print_with_color(f"\n\tto_date : {to_date}\n", 28)
            crawlerObj.to_date = to_date
            break
    else : pass

    # argparser로 druglist path를 받지 못했을 경우
    druglist_path = kwargs["druglist_path"]
    if not crawlerObj.check_druglist_path(druglist_path, first_check=True):
        # druglist_path 사용자 입력
        print_with_color("\n2. Enter the Excel file path to set up the drugs.", 33)
        print_with_color("   If you don't have an Excel file, Enter Empty input", 33)
        print_with_color("   input form - (ex : data/drugs_20220301.xlsx)", 33)
        while True:
            druglist_path = input(":")

            if not crawlerObj.check_druglist_path(druglist_path):
                continue
            else : pass

            # crawling obj에 drugpath_list attribute update
            print_with_color(f"\n\tdruglist_path : {druglist_path}\n", 28)
            crawlerObj.druglist_path = druglist_path
            break
    else : pass

    columns = kwargs["columns"]
    drugs = kwargs["drugs"]

    # druglist_path는 있지만 columns는 없을 경우
    if druglist_path and not columns:
        print_with_color("*" * 50, 33)
        print()
        print_with_color("\n2-1. Enter the columns to set up the drugs in excel file.", 33)
        print_with_color("     If you're done typing, Enter Empty input", 33)
        print_with_color("     delete last input - ;;b , view columns list - ;;li", 33)
        while True:
            column = input(":")

            # ;;b는 입력받은 column 중 가장 최근 값을 삭제
            if column == ";;b":
                # 만약 입력을 받지 않고 삭제하려고 한다면 다시시도
                if not columns:
                    print_with_color(f"\t*** columns list is empty ***", 9)
                    continue
                else:
                    print_with_color(f"{columns.pop()} is deleted", 9)
                    continue

            # ;;li는 지금까지 입력받은 column들을 print하는 명령
            elif column == ";;li":
                print_with_color("\tcolumns list", 28)
                print_with_color(f"\t{columns}", 28)
                continue

            # 빈 입력값을 받았다면 입력을 완료했다고 하여 columns attribute update & 입력 종료
            elif not column:
                print_with_color(f"\n\tcolumns : {columns}\n", 28)
                crawlerObj.columns = columns
                break

            # 일반적인 column name을 입력하려는 경우 columns list에 append
            else:
                columns.append(column)

    # 엑셀파일로 읽지 않는다면 타이핑으로 drug를 하나씩 입력
    elif (not druglist_path) and (not columns) and (not drugs):
        print_with_color("*" * 50, 33)
        print_with_color("\n2-1. Enter the drug to set up the drugs.", 33)
        print_with_color("     If you're done typing, Enter Empty input", 33)
        print_with_color("     delete last input - ;;b , view drugs list - ;;li", 33)
        while True:
            drug = input(":")

            # ;;b는 입력받은 drug 중 가장 최근 값을 삭제
            if drug == ";;b":
                # 만약 입력을 받지 않고 삭제하려고 한다면 다시시도
                if not drugs:
                    print_with_color(f"\t*** drugs list is empty ***", 9)
                    continue
                else:
                    print_with_color(f"{drugs.pop()} is deleted", 9)
                    continue

            # ;;li는 지금까지 입력받은 drug들인 drugs list를 print하는 명령
            elif drug == ";;li":
                print_with_color("\tdrugs list", 28)
                print_with_color(f"\t{drugs}", 28)
                continue

            # 빈 입력값을 받았다면 입력을 완료했다고 하여 drugs attribute update & 입력 종료
            elif not drug:
                if not drugs:
                    print_with_color(f"\t*** drugs list is empty ***", 9)
                    continue
                else:
                    print_with_color(f"\n\tdrugs : {drugs}\n", 28)
                    crawlerObj.drugs = drugs
                    break

            # 일반적인 drug name을 입력하려는 경우 drugs list에 append
            else:
                drugs.append(drug)

    else : pass

    '''
    
        drugs & brands setting 

    '''

    drg_dbcol = crawlerObj.get_mongoDB_access(host=crawlerObj.mongodb_brand_host ,
                                              db_name=crawlerObj.mongodb_brand_db,
                                              collection_name=crawlerObj.mongodb_brand_col)

    regx_str = crawlerObj.get_brands_find_args_with_fisrtchar(crawlerObj.brands_find_re_charlist)

    crawlerObj.set_drugs_brands_from_mongoDB(mongoDB_col=drg_dbcol,
                                            find_args=regx_str if crawlerObj.brands_find_re_charlist else {},
                                            find_kwargs={"_id": 0, "update_time": 0})
    '''

        crawling & preprocessing & sui_classifier & mongoDB & doccano

    '''


    # crawling - http GET requests

    crawlerObj.crawling_by_brand_and_drug()



    '''

        make histogram

    '''

    # sui_classifier 값들을 정렬 후 리스트화 , 통계적 수치들을 딕셔너리 화
    # suiclasifier_list, statistics_dict = crawlerObj.make_histogram_data_and_StatisticsDict(biger_suiclsf=histogram_sui_clasifier_limit)

    # crawling 정보 - from_date , to_date , source 저장
    # crawling_inform = crawlerObj.make_crawling_inform()

    # 빈 입력 시 위의 포맷으로 이름 생성
    # if histogram_name == "":
       # from datetime import datetime

       # now = datetime.now()
       # histogram_name = f'histogram/{crawlerObj.source}_histogram{now.strftime("%Y%m%d")}_1.png'
       # file_num = 1
       # while True:
       #     if os.path.isfile(histogram_name):
       #         file_num += 1
       #         histogram_name = f'histogram/{crawlerObj.source}_histogram{now.strftime("%Y%m%d")}_{file_num}.png'
       #         continue
       #     break

    # png를 제외한 이름이면 확장자인 .png 추가후 이름 생성
    #elif (len(histogram_name) > 5 and histogram_name[-4:] != ".png") or len(histogram_name) < 4:
    #    histogram_name = histogram_name + ".png"
    #else : pass

    # sui_classifier를 토대로한 히스토그램 생성 - 통계적 수치들과 크롤링 정보들은 그래프 내부 텍스트로 입력
    #print(f"making histogram {histogram_name}...")
    #crawlerObj.make_histogram(suiclasifier_list, crawling_inform, statistics_dict, histogram_name)

    '''

        update mongoDB

    '''

    # 크롤링 데이터 mongoDB로 저장
    #crawlerObj.update_to_MongoDB()

    '''

        export to excel file

    '''

    # excel로 저장한다는
    #if is_excel_save:
    #    crawlerObj.export_excelfile(host=crawlerObj.mongodb_host, collection_name=crawlerObj.mongodb_col,
    #                     documents_name=crawlerObj.mongodb_doc, limit=excelDB_limit,
    #                     search_q=excelDB_query, file_name=excelDB_filename,
    #                     sentences_keys=excelDB_sentences_keys)
    #else : pass

if __name__ == "__main__":
    import configparser as cparser
    import os.path
    import json
    import argparse
    from argparse import RawTextHelpFormatter

    # using argparse

    """

    Command Prompt

    """

    parser = argparse.ArgumentParser(description="\tNBDMRcrawler - TweetCrawler\n\n"
                                                 "\tPython crawler to collect drug-related data from twitter.\n"
                                                 "\tYou can crawling tweets\n"
                                                 "\tSupport for saving crawling data in MongoDB \n"
                                                 "\tSupport for saving crawling data(tweets) to excel files \n"
                                                 "\t excel files is stored in labels directory\n"
                                                 "\tSupport Histogram including suicide clasifier and statistics data\n"
                                                 "\t histogram files is stored in histogram directory\n\n",
                                     formatter_class=RawTextHelpFormatter)

    parser.add_argument('--start_date' , '-sd', dest="from_date", type=str,
                        help="The date (YYYY/MM/DD) after which the publication was published\n ",
                        default="")

    parser.add_argument('--end_date' , '-ed', dest="to_date", type=str,
                        help="The date (YYYY/MM/DD) before which the publication was published\n ",
                        default="")

    parser.add_argument('--drugs','-drgs', dest="drugs", nargs='+',
                        help="Array of drug names. \n"
                             "if drgs argument has value, druglist_path & columns does not work.\n"
                             "only this argument will provide drug list to crawler.\n ",
                        default=[])

    parser.add_argument('--columns','-clms', dest='columns', nargs='+',
                        help='columns for crawling drug list in druglist_path\n ',
                        default=[])

    parser.add_argument('--druglist_path','-drgsp', dest="druglist_path", type=str,
                        help='Path for a excel with drug list\n ',
                        default="")

    parser.add_argument('--config_path','-cnfgp', dest="config_path", type=str,
                        help="config file path\n ",
                        default="config/config.ini")

    parser.add_argument('--histogram_name','-hstn', dest="histogram_name", type=str,
                        help="png file name for save histogram.\n"
                             "if argument empty , file name will be this form (ex: tweet_histogram_20220301.png)\n ",
                        default="")

    parser.add_argument('--histogram_sui_clasifier_limit', '-hstlm', dest="histogram_sui_clasifier_limit", type=float,
                        help="If the histogram is less than the specified number in the list to create it,\n"
                             "do not reflect it in the histogram.\n ",
                        default=None)

    parser.add_argument('--is_excel_save','-xlsxsv', dest="is_excel_save",
                        help="if you want to save excel file,\n"
                             "just add 'xlsxsv' string in prompt parsing line.\n"
                             "no argument xlsxsv, not saved excel file.\n ",
                        action="store_true")

    parser.add_argument('--excel_name','-xlsxn', dest="excel_name", type=str,
                        help="name to save excel file with crawling data.\n"
                             "if argument empty , save excel file with formatting name like 'Tweet20220209_to_label.xls\n ",
                        default=None)

    parser.add_argument('--excel_limit','-xlsxlm', dest="excel_limit", type=int,
                        help="the number of data in MongoDB to be made into Excel,\n "
                             "and all data are used if argument empty.\n ",
                        default=None)

    parser.add_argument('--excel_query','-xlsxsq', dest="excel_query", type=int,
                        help="search Query of MongoDB data to be made into Excel,\n "
                             "and all data are used if argument empty.\n ",
                        default=None)

    crawling_options = vars(parser.parse_args())

    # config_path
    config_path = crawling_options["config_path"]

    # config_file_path 입력이 잘못되었다면 사용자 입력으로 다시 받기
    if not os.path.isfile(config_path):
        print_with_color(f"\n** there is no file in {config_path} **", 9)
        print_with_color("\nEnter the config file path to set up crawler", 33)
        print_with_color("config file form is .ini", 33)
        print_with_color("input form - (ex : config/config.ini)", 33)
        while True:
            config_path = input(":")

            # 입력받은 path에 파일이 없으면 다시시도
            if not os.path.isfile(config_path):
                print_with_color(f"\t*** there is no file in {config_path} ***", 9)
                continue

            print_with_color(f"\n\tconfig_path : {config_path}", 28)
            break
    else : pass
    
    # config file read start
    config = cparser.ConfigParser()
    config.read(config_path)

    # config 파일 값 읽기
    for section in config.sections():
        for item in config.items(section):
            if item[1] in ["True", "False"]:
                crawling_options[item[0]] = config.getboolean(section, item[0])
            elif item[1] == "None":
                crawling_options[item[0]] = None
            elif item[1].isdigit():
                crawling_options[item[0]] = int(item[1])
            elif len(item[1]) > 1 and item[1][0] == "[" and item[1][-1] == "]":
                replace_single_quot = item[1].replace("\'", "\"")
                crawling_options[item[0]] = json.loads(replace_single_quot)
            else:
                crawling_options[item[0]] = item[1]

    # tweet crawler 이기 때문에 source는 tweet으로 고정
    crawling_options["source"] = "tweet"

    main(**crawling_options)

